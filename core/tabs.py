"""
tabs.py — workbook construction for the analytics engine.

Reproduces the uniform Ziggurat performance-workbook format (verified against a real
ChefSteps May-2026 deliverable). The LAYOUT is identical for every client; only flag/split
inputs (trigger word, Trial Reel copy) vary, and those arrive already-applied in the dataframes.

Tabs:
  build_feed_tab     Facebook, Instagram, and Trial Reels (same shape; CTA columns differ)
  build_stories_tab  Stories (purple month divider, header on row 2, data from row 3)
  build_footnotes    format-breakdown tables (data) + insight shells (model fills)
  build_summary      combined snapshot (data) + executive shells (model fills)

Header row is row 1 for feed tabs; data starts row 2. Stories offsets by one for its divider.
"""
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ---- palette (verified from the real workbook) ------------------------------------------------
C = {
    "post_hdr": "2C2C2C", "metric_hdr": "2A4A44", "total_hdr": "8B2020",
    "score_hdr": "3D4B55", "summary_hdr": "2C4A8C", "stories_hdr": "5B3A7A", "accent": "FC684E",
    "rank1": "1AAD4A", "top": "D6F0D6", "bottom": "FFD6D6",
    "top_title": "006100", "bottom_title": "8B0000",
    "alt": "FAF9F7", "white": "FFFFFF",
    "score_bg": "FFF3F0", "total_bg": "FFE8E4",
    "label_bg": "EFE9E4", "caption_bg": "F0F4F8", "trigger_text": "2A4A44",
}
THIN = Side(style="thin", color="FFD0D0D0")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
ARIAL = "Arial"


def _fill(hex_):
    return PatternFill("solid", fgColor="FF" + hex_ if len(hex_) == 6 else hex_)


def _hdr_cell(ws, r, c, text, bg):
    cell = ws.cell(r, c, text)
    cell.font = Font(name=ARIAL, bold=True, color="FFFFFFFF", size=11)
    cell.fill = _fill(bg)
    cell.alignment = Alignment(vertical="center", wrap_text=True)
    cell.border = BORDER
    return cell


# ----------------------------------------------------------------------------------------------
# Feed tab — Facebook / Instagram / Trial Reels
# ----------------------------------------------------------------------------------------------
def build_feed_tab(wb, df, cfg, month, kind, title=None, top_note=None):
    """
    kind: "facebook" | "instagram"  (Trial Reels uses "instagram" with a top_note)
    df:   scored dataframe, already sorted by Overall Rank asc, with Score:/Total Score/Overall Rank
          and (IG) Trigger Word/Link in Bio/No CTA or (FB) Link in Comments columns present.
    """
    ws = wb.create_sheet(title=title or f"{kind.title()} — {month}")
    metrics = cfg.metrics.get(kind, [])
    n = len(df)
    top_n, bottom_n = 5, 5
    top_thresh = df["Total Score"].nlargest(top_n).min() if n else 0
    bottom_thresh = df["Total Score"].nsmallest(bottom_n).max() if n else 0

    # --- column plan -------------------------------------------------------------------------
    cols = [("#", "num"), ("Post Title", "title"), ("Format", "format")]
    if kind == "facebook":
        cols += [("Link in Comments", "cta")]
    else:
        cols += [("Trigger Word", "cta"), ("Link in Bio", "cta"), ("No CTA", "cta")]
    cols += [("Date Published", "date")]
    raw_start = len(cols) + 1                          # first raw-metric column (1-indexed)
    cols += [(m, "raw") for m in metrics]
    cols += [(f"Score: {m}", "score") for m in metrics]
    cols += [("Total Score", "total"), ("Overall Rank", "rank"),
             ("Performance Note", "note"), ("Permalink", "permalink"), ("Full Caption", "caption")]
    widths = {"num": 4, "title": 46, "format": 14, "cta": 13, "date": 14, "raw": 11,
              "score": 9, "total": 12, "rank": 11, "note": 42, "permalink": 30, "caption": 3}

    # --- headers (row 1) ---------------------------------------------------------------------
    for i, (label, kind_) in enumerate(cols, start=1):
        if kind_ == "raw":
            bg = C["metric_hdr"]
        elif kind_ == "score":
            bg = C["score_hdr"]
        elif label in ("Total Score", "Overall Rank"):
            bg = C["total_hdr"]
        elif label == "Full Caption":
            bg = "5C3D2E"                      # warm brown, distinct caption header
        else:
            bg = C["post_hdr"]                 # #, Title, Format, CTA, Date, Note, Permalink
        _hdr_cell(ws, 1, i, label, bg)
        ws.column_dimensions[get_column_letter(i)].width = widths[kind_]

    # --- data rows (row 2+) ------------------------------------------------------------------
    for ridx, (_, post) in enumerate(df.iterrows()):
        r = ridx + 2
        rank = int(post["Overall Rank"])
        total = int(post["Total Score"])
        in_top = ridx < top_n                       # position-based: first N rows are the top band
        in_bottom = ridx >= n - bottom_n            # last N rows are the bottom band
        base = C["top"] if in_top else C["bottom"] if in_bottom else (C["alt"] if ridx % 2 else C["white"])

        for i, (label, kind_) in enumerate(cols, start=1):
            cell = ws.cell(r, i)
            cell.font = Font(name=ARIAL, size=10)
            cell.border = BORDER
            cell.fill = _fill(base)
            cell.alignment = Alignment(vertical="center", horizontal="left")

            if kind_ == "num":
                cell.value = rank
            elif kind_ == "title":
                cell.value = _title(post)
                col = C["top_title"] if in_top else C["bottom_title"] if in_bottom else "000000"
                cell.font = Font(name=ARIAL, size=10, bold=in_top or in_bottom, color="FF" + col)
            elif kind_ == "format":
                cell.value = _format_label(post.get("Post type", ""), kind)
            elif label == "Link in Comments":
                cell.value = "✓" if post.get("Link in Comments") else None
            elif label == "Trigger Word":
                tw = post.get("Trigger Word") or ""
                cell.value = tw or None
                if tw:
                    cell.fill = _fill(C["top"])
                    cell.font = Font(name=ARIAL, size=10, bold=True, color="FF"+C["trigger_text"])
                else:
                    cell.fill = _fill("FAFAFA")      # blank trigger cell sits on near-white, not row tint
            elif label == "Link in Bio":
                cell.value = "✓" if post.get("Link in Bio") else None
            elif label == "No CTA":
                cell.value = "✓" if post.get("No CTA") else None
            elif kind_ == "date":
                cell.value = _date(post.get("Publish time"))
            elif kind_ == "raw":
                cell.value = _num(post.get(label))
            elif kind_ == "score":
                cell.value = int(post.get(label, 0))
                cell.fill = _fill(C["score_bg"])      # score cells always keep their own tint
            elif label == "Total Score":
                cell.value = total
                cell.fill = _fill(C["total_bg"])      # total cell always keeps its own tint
            elif label == "Overall Rank":
                cell.value = rank
                cell.alignment = Alignment(horizontal="center", vertical="center")
                if rank == 1:
                    cell.fill = _fill(C["rank1"]); cell.font = Font(name=ARIAL, size=10, bold=True, color="FFFFFFFF")
                elif in_top:
                    cell.fill = _fill(C["top"]); cell.font = Font(name=ARIAL, size=10, bold=True)
                elif in_bottom:
                    cell.fill = _fill(C["bottom"]); cell.font = Font(name=ARIAL, size=10, bold=True)
            elif label == "Performance Note":
                note = _perf_note(post, df, metrics, top_thresh, bottom_thresh, top_n, bottom_n, n, in_top, in_bottom)
                cell.value = note
                cell.alignment = Alignment(vertical="top", horizontal="left", wrap_text=True)
            elif label == "Permalink":
                url = post.get("Permalink")
                if isinstance(url, str) and url.startswith("http"):
                    cell.value = url; cell.hyperlink = url
                    cell.font = Font(name=ARIAL, size=10, color="FF2C4A8C", underline="single")
            elif label == "Full Caption":
                cell.value = post.get("Description")
                cell.fill = _fill(C["caption_bg"])
                cell.alignment = Alignment(vertical="center", horizontal="left", wrap_text=True)

    ws.freeze_panes = f"{get_column_letter(raw_start)}2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(cols))}1"
    if top_note:
        ws.insert_rows(1)
        ws.cell(1, 1, top_note).font = Font(name=ARIAL, italic=True, size=10)
    return ws


def build_empty_special_tab(wb, name, month, note):
    ws = wb.create_sheet(title=f"{name} — {month}")
    ws.cell(1, 1, f"{name} are scored independently and excluded from the main Instagram performance data.")
    ws.cell(2, 1, note)
    return ws


# ----------------------------------------------------------------------------------------------
# helpers
# ----------------------------------------------------------------------------------------------
def _title(post, limit=65):
    import pandas as pd
    src = post.get("Title") or post.get("Description") or ""
    if src is None or (isinstance(src, float) and pd.isna(src)):
        src = ""
    line = str(src).strip().split("\n")[0]
    if not line or line.lower() == "nan":
        return "(no description)"
    return line if len(line) <= limit else line[:limit] + "…"


def _format_label(post_type, kind):
    pt = str(post_type or "")
    if kind == "facebook":
        return "Video/Reel" if "ideo" in pt else "Photo/Carousel"
    return {"IG reel": "Reel", "IG carousel": "Carousel", "IG image": "Image"}.get(pt, pt)


def _date(val):
    import pandas as pd
    t = pd.to_datetime(val, errors="coerce")
    return t.strftime("%m/%d/%Y %H:%M") if pd.notna(t) else (str(val) if val else "")


def _num(v):
    import pandas as pd
    return int(v) if pd.notna(v) and float(v) == int(float(v)) else (float(v) if pd.notna(v) else 0)


def _perf_note(post, df, metrics, top_thresh, bottom_thresh, top_n, bottom_n, n, in_top, in_bottom):
    """Populated for every middle-tier post (not top 5, not bottom 5), matching the real workbook:
    distance to the nearer threshold + weakest metric, with a borderline flag only when within 15."""
    if in_top or in_bottom:
        return None
    total = int(post["Total Score"])
    to_top = top_thresh - total
    to_bottom = total - bottom_thresh
    if to_top <= to_bottom:
        prox = f"{to_top} pts from the top 5 threshold."
        borderline = 0 < to_top <= 15
    else:
        prox = f"{to_bottom} pts from the bottom 5 threshold."
        borderline = 0 < to_bottom <= 15
    if borderline:
        prox += " (Within 15 — borderline.)"
    weak = min(metrics, key=lambda mm: post.get(f"Score: {mm}", 9999))
    weak_rank = int(post.get(f"Score: {weak}", 0))
    note = f"{prox} Lowest ranked metric: {weak} (ranked {weak_rank}/{n})."

    # CTA observation — flags a missed/partial conversion setup, matching the real workbook.
    if "Link in Comments" in post.index:  # Facebook
        if not post.get("Link in Comments"):
            note += " No link-in-comments CTA — broadcast post only."
    else:  # Instagram
        if post.get("No CTA"):
            note += " No CTA — missed conversion opportunity."
        elif post.get("Link in Bio") and not (post.get("Trigger Word") or ""):
            note += " Link in bio only — no trigger word to drive the auto-DM."
    return note


# ----------------------------------------------------------------------------------------------
# Stories tab — month divider (row 1), headers (row 2), data (row 3+)
# ----------------------------------------------------------------------------------------------
def build_stories_tab(wb, df, cfg, month):
    ws = wb.create_sheet(title=f"Stories — {month}")
    metrics = [m for m in cfg.metrics.get("stories", []) if f"Score: {m}" in df.columns]
    n = len(df)
    top_n, bottom_n = 10, 10
    top_thresh = df["Total Score"].nlargest(top_n).min() if n else 0
    bottom_thresh = df["Total Score"].nsmallest(bottom_n).max() if n else 0

    cols = [("#", "num", 4), ("Month", "format", 12), ("Story Description", "title", 46),
            ("Date", "date", 16), ("Duration (sec)", "raw", 12)]
    raw_start = len(cols) + 1
    cols += [(m, "raw", 11) for m in metrics]
    cols += [(f"Score: {m}", "score", 9) for m in metrics]
    cols += [("Total Score", "total", 12), ("Overall Rank", "rank", 11),
             ("Permalink", "permalink", 30), ("Full Caption", "caption", 3)]

    # divider row 1
    ws.cell(1, 1, f"{month} — {n} stories")
    for c in range(1, len(cols) + 1):
        ws.cell(1, c).fill = _fill(C["stories_hdr"])
        ws.cell(1, c).font = Font(name=ARIAL, bold=True, color="FFFFFFFF", size=11)

    for i, (label, kind_, w) in enumerate(cols, start=1):
        if kind_ == "raw" and label != "Duration (sec)":
            bg = C["metric_hdr"]
        elif kind_ == "score":
            bg = C["score_hdr"]
        elif label in ("Total Score", "Overall Rank"):
            bg = C["total_hdr"]
        elif label == "Full Caption":
            bg = "5C3D2E"
        else:
            bg = C["post_hdr"]               # #, Month, Description, Date, Duration, Permalink
        _hdr_cell(ws, 2, i, label, bg)
        ws.column_dimensions[get_column_letter(i)].width = w

    for ridx, (_, post) in enumerate(df.iterrows()):
        r = ridx + 3
        rank = int(post["Overall Rank"]); total = int(post["Total Score"])
        in_top = ridx < top_n; in_bottom = ridx >= n - bottom_n
        base = C["top"] if in_top else C["bottom"] if in_bottom else (C["alt"] if ridx % 2 else C["white"])
        for i, (label, kind_, w) in enumerate(cols, start=1):
            cell = ws.cell(r, i); cell.font = Font(name=ARIAL, size=10)
            cell.border = BORDER; cell.fill = _fill(base)
            cell.alignment = Alignment(vertical="center", horizontal="left")
            if kind_ == "num": cell.value = rank
            elif label == "Month": cell.value = month
            elif kind_ == "title": cell.value = _title({"Description": post.get("Description")}, limit=60)
            elif kind_ == "date": cell.value = _date(post.get("Publish time"))
            elif kind_ == "raw": cell.value = _num(post.get(label))
            elif kind_ == "score":
                cell.value = int(post.get(label, 0))
                cell.fill = _fill(C["score_bg"])
            elif label == "Total Score":
                cell.value = total
                cell.fill = _fill(C["total_bg"])
            elif label == "Overall Rank":
                cell.value = rank; cell.alignment = Alignment(horizontal="center", vertical="center")
                if rank == 1: cell.fill = _fill(C["rank1"]); cell.font = Font(name=ARIAL, size=10, bold=True, color="FFFFFFFF")
            elif label == "Permalink":
                url = post.get("Permalink")
                if isinstance(url, str) and url.startswith("http"):
                    cell.value = url; cell.hyperlink = url
                    cell.font = Font(name=ARIAL, size=10, color="FF2C4A8C", underline="single")
            elif label == "Full Caption":
                cell.value = post.get("Description"); cell.fill = _fill(C["caption_bg"])
                cell.alignment = Alignment(vertical="center", wrap_text=True)
    ws.freeze_panes = f"{get_column_letter(raw_start)}3"
    return ws


# ----------------------------------------------------------------------------------------------
# Format-breakdown computation (deterministic — the script hands these tables to the model)
# ----------------------------------------------------------------------------------------------
def ig_breakdown(df, metrics):
    """IG format table: Reels | Carousels | Images. Curated row list matching the standard format."""
    groups = {"Reels": df[df["Post type"] == "IG reel"],
              "Carousels": df[df["Post type"] == "IG carousel"],
              "Images": df[df["Post type"] == "IG image"]}
    sums = ["Views", "Reach", "Likes", "Comments", "Shares", "Saves", "Follows"]
    avgs = ["Views", "Saves", "Comments"]
    return _build_rows(df, groups, sums, avgs, with_total=False), False


def fb_breakdown(df, metrics):
    """FB format table: Video/Reel | Photo/Carousel | Platform Total. Uses Link clicks (labelled
    'Link Clicks'), and a trailing 'Link in Comments posts' count row."""
    vids = df[df["Post type"].str.contains("ideo", na=False)]
    photos = df[~df["Post type"].str.contains("ideo", na=False)]
    groups = {"Video/Reel": vids, "Photo/Carousel": photos}
    sums = [("Views", "Views"), ("Reach", "Reach"), ("Reactions", "Reactions"),
            ("Comments", "Comments"), ("Shares", "Shares"), ("Link Clicks", "Link clicks")]
    avgs = [("Avg Views/post", "Views"), ("Avg Reactions/post", "Reactions")]
    rows = [("Posts", {g: len(d) for g, d in groups.items()}, len(df), "count")]
    import pandas as pd
    for label, col in sums:
        if col not in df.columns: continue
        rows.append((label, {g: pd.to_numeric(d[col], errors="coerce").fillna(0).sum() for g, d in groups.items()},
                     pd.to_numeric(df[col], errors="coerce").fillna(0).sum(), "sum"))
    for label, col in avgs:
        if col not in df.columns: continue
        rows.append((label, {g: (pd.to_numeric(d[col], errors="coerce").fillna(0).mean() if len(d) else 0) for g, d in groups.items()},
                     pd.to_numeric(df[col], errors="coerce").fillna(0).mean() if len(df) else 0, "avg"))
    if "Link in Comments" in df.columns:
        rows.append(("Link in Comments posts",
                     {g: int(d["Link in Comments"].sum()) for g, d in groups.items()},
                     int(df["Link in Comments"].sum()), "count"))
    return rows, True


def _build_rows(df, groups, sums, avgs, with_total):
    import pandas as pd
    rows = [("Posts", {g: len(d) for g, d in groups.items()}, len(df), "count")]
    for col in sums:
        if col not in df.columns: continue
        rows.append((col, {g: pd.to_numeric(d[col], errors="coerce").fillna(0).sum() for g, d in groups.items()},
                     pd.to_numeric(df[col], errors="coerce").fillna(0).sum(), "sum"))
    for col in avgs:
        if col not in df.columns: continue
        rows.append((f"Avg {col}/post",
                     {g: (pd.to_numeric(d[col], errors="coerce").fillna(0).mean() if len(d) else 0) for g, d in groups.items()},
                     None, "avg"))
    return rows


def _section_header(ws, r, text, bg, span=2):
    for c in range(1, span + 1):
        ws.cell(r, c).fill = _fill(bg)
        ws.cell(r, c).font = Font(name=ARIAL, bold=True, color="FFFFFFFF", size=11)
    ws.cell(r, 1, text)


def _pct(part, total):
    return f"{part:,.0f}  ({part / total * 100:.0f}%)" if total else f"{part:,.0f}  (0%)"


def build_footnotes(wb, cfg, month, ig_df=None, fb_df=None, st_df=None, insights=None):
    """Data tables are computed and written; insight sections are shells unless `insights` given."""
    ws = wb.create_sheet(title="Footnotes & Insights")
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 60
    ws.cell(1, 1, f"{cfg.client_name} Performance Footnotes & Insights — {month}").font = Font(name=ARIAL, bold=True, size=13)
    r = 3

    if ig_df is not None:
        rows, _ = ig_breakdown(ig_df, cfg.metrics.get("instagram", []))
        _section_header(ws, r, "Instagram Format Breakdown", C["metric_hdr"], span=4); r += 1
        for c, h in enumerate(["Metric", "Reels", "Carousels", "Images"], start=1):
            _hdr_cell(ws, r, c, h, C["score_hdr"])
        r += 1
        for label, vals, _tot, kind in rows:
            ws.cell(r, 1, label).fill = _fill(C["label_bg"]); ws.cell(r, 1).font = Font(name=ARIAL, bold=True, size=12)
            total = sum(vals.values())
            for c, (g, v) in enumerate(vals.items(), start=2):
                txt = str(int(v)) if kind == "count" else (f"{v:.1f}" if kind == "avg" else _pct(v, total))
                ws.cell(r, c, txt).fill = _fill(C["label_bg"])
            r += 1
        r += 1

    if fb_df is not None:
        rows, _ = fb_breakdown(fb_df, cfg.metrics.get("facebook", []))
        _section_header(ws, r, "Facebook Format Breakdown", C["summary_hdr"], span=4); r += 1
        for c, h in enumerate(["Metric", "Video/Reel", "Photo/Carousel", "Platform Total"], start=1):
            _hdr_cell(ws, r, c, h, C["summary_hdr"])
        r += 1
        for label, vals, tot, kind in rows:
            ws.cell(r, 1, label).fill = _fill(C["label_bg"]); ws.cell(r, 1).font = Font(name=ARIAL, bold=True, size=12)
            grp_total = sum(vals.values())
            for c, (g, v) in enumerate(vals.items(), start=2):
                txt = str(int(v)) if kind == "count" else (f"{v:.1f}" if kind == "avg" else _pct(v, grp_total))
                ws.cell(r, c, txt).fill = _fill(C["label_bg"])
            # Platform Total column
            if tot is not None:
                tot_txt = str(int(tot)) if kind == "count" else (f"{tot:.1f}" if kind == "avg" else f"{tot:,.0f}")
                ws.cell(r, 4, tot_txt).fill = _fill(C["label_bg"])
            r += 1
        r += 1

    if st_df is not None:
        import pandas as pd
        _section_header(ws, r, "Stories Breakdown", C["stories_hdr"], span=3); r += 1
        for c, h in enumerate(["Metric", month, "Period Total"], start=1):
            _hdr_cell(ws, r, c, h, C["stories_hdr"])
        r += 1
        n_st = len(st_df)
        def tot(col):
            return pd.to_numeric(st_df[col], errors="coerce").fillna(0).sum() if col in st_df.columns else 0
        st_rows = [("Stories published", n_st, "count"),
                   ("Total Views", tot("Views"), "sum"), ("Total Reach", tot("Reach"), "sum"),
                   ("Total Likes", tot("Likes"), "sum"), ("Total Replies", tot("Replies"), "sum"),
                   ("Total Link clicks", tot("Link clicks"), "sum"),
                   ("Total Profile visits", tot("Profile visits"), "sum"),
                   ("Avg Views/story", tot("Views") / n_st if n_st else 0, "avg"),
                   ("Avg Reach/story", tot("Reach") / n_st if n_st else 0, "avg"),
                   ("Avg Link clicks/story", tot("Link clicks") / n_st if n_st else 0, "avg")]
        for label, v, kind in st_rows:
            ws.cell(r, 1, label).fill = _fill(C["label_bg"]); ws.cell(r, 1).font = Font(name=ARIAL, bold=True, size=12)
            txt = str(int(v)) if kind == "count" else (f"{v:.1f}" if kind == "avg" else f"{v:,.0f}")
            ws.cell(r, 2, txt).fill = _fill(C["label_bg"])
            ws.cell(r, 3, txt).fill = _fill(C["label_bg"])   # single-month period: total == month
            r += 1
        r += 1

    # Insight shells — the model fills these (or `insights` if pre-written)
    for title, bg, default_n in [("What Worked", C["post_hdr"], 4),
                                  ("What Didn't Work", C["total_hdr"], 3),
                                  ("What We Do About It", C["accent"], 4)]:
        _section_header(ws, r, title, bg); r += 1
        items = (insights or {}).get(title, [])
        if items:
            for label, body in items:
                ws.cell(r, 1, label).font = Font(name=ARIAL, bold=True, size=12)
                ws.cell(r, 2, body).font = Font(name=ARIAL, size=11)
                ws.cell(r, 2).alignment = Alignment(wrap_text=True, vertical="top")
                ws.row_dimensions[r].height = 78
                r += 1
        else:
            ws.cell(r, 1, f"[{title}: {default_n} entries — to be written]").font = Font(name=ARIAL, italic=True, size=11, color="FF999999")
            r += 1
        r += 1
    return ws


def build_summary(wb, cfg, month, ig_df=None, fb_df=None, st_df=None, insights=None, extra_note=None):
    import pandas as pd
    ws = wb.create_sheet(title="Summary & Strategy")
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 60
    ws.cell(1, 1, f"{cfg.client_name} Social — Executive Summary — {month}").font = Font(name=ARIAL, bold=True, size=13)
    r = 3

    _section_header(ws, r, "Combined Platform Snapshot", C["post_hdr"]); r += 1
    for c, h in enumerate(["Instagram", "Facebook", "Combined"], start=2):
        _hdr_cell(ws, r, c, h, C["metric_hdr"])
    r += 1

    def s(df, col):
        return pd.to_numeric(df[col], errors="coerce").fillna(0).sum() if df is not None and col in df.columns else 0
    ig_posts = len(ig_df) if ig_df is not None else 0
    fb_posts = len(fb_df) if fb_df is not None else 0
    snap = [
        ("Posts published", ig_posts, fb_posts, ig_posts + fb_posts, "plain"),
        ("Total Views", s(ig_df, "Views"), s(fb_df, "Views"), s(ig_df, "Views") + s(fb_df, "Views"), "comma"),
        ("Avg Views/post", s(ig_df, "Views") / ig_posts if ig_posts else 0,
         s(fb_df, "Views") / fb_posts if fb_posts else 0,
         (s(ig_df, "Views") + s(fb_df, "Views")) / (ig_posts + fb_posts) if (ig_posts + fb_posts) else 0, "plain"),
        ("Total Saves (IG)", s(ig_df, "Saves"), "—", s(ig_df, "Saves"), "comma"),
        ("Total Follows gained (IG)", s(ig_df, "Follows"), "—", s(ig_df, "Follows"), "comma"),
        ("Total Comments", s(ig_df, "Comments"), s(fb_df, "Comments"), s(ig_df, "Comments") + s(fb_df, "Comments"), "comma"),
    ]
    for label, ig_v, fb_v, comb, fmt in snap:
        ws.cell(r, 1, label).fill = _fill(C["label_bg"]); ws.cell(r, 1).font = Font(name=ARIAL, bold=True, size=12)
        for c, v in enumerate([ig_v, fb_v, comb], start=2):
            txt = v if isinstance(v, str) else (f"{v:,.0f}" if fmt == "comma" else f"{v:.0f}")
            ws.cell(r, c, txt).fill = _fill(C["label_bg"])
        r += 1
    r += 1

    _section_header(ws, r, "Executive View", C["summary_hdr"]); r += 1
    lenses = ["1. The creative formula", "2. CTA and conversion mechanics",
              "3. Platform differentiation", "4. Volume vs quality", "5. Strategic context"]
    ev = (insights or {}).get("Executive View", [])
    for i, lens in enumerate(lenses):
        ws.cell(r, 1, lens).font = Font(name=ARIAL, bold=True, size=12)
        body = ev[i][1] if i < len(ev) else "[to be written from config strategy + this month's data]"
        ws.cell(r, 2, body).font = Font(name=ARIAL, size=11, italic=(i >= len(ev)),
                                        color=("999999" if i >= len(ev) else None))
        ws.cell(r, 2).alignment = Alignment(wrap_text=True, vertical="top")
        ws.row_dimensions[r].height = 90
        r += 1
    r += 1

    _section_header(ws, r, "Recommended Focus Areas", C["accent"]); r += 1
    fa = (insights or {}).get("Recommended Focus Areas", [])
    if fa:
        for label, body in fa:
            ws.cell(r, 1, label).font = Font(name=ARIAL, bold=True, size=12)
            ws.cell(r, 2, body).font = Font(name=ARIAL, size=11)
            ws.cell(r, 2).alignment = Alignment(wrap_text=True, vertical="top")
            ws.row_dimensions[r].height = 78
            r += 1
    else:
        ws.cell(r, 1, "[3–4 numbered focus areas — to be written]").font = Font(name=ARIAL, italic=True, size=11, color="FF999999")
        r += 1
    if extra_note:
        r += 1
        ws.cell(r, 1, extra_note).font = Font(name=ARIAL, italic=True, size=10)
    return ws
